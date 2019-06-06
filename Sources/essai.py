from openpyxl import Workbook
from openpyxl import load_workbook

verbose = 10

"""
Class pour la gestion d'une feuille de résultat
"""
class CLASS_resultSheet:
    class CLASS_indexCol:
        class CLASS_indexColData:
            def __init__(self, refCol, col):
                self.refCol = refCol
                self.col = col
        def __init__(self):
            self.list = []
        def append(self, refCol, col):
            self.list.append(self.CLASS_indexColData(refCol, col))
            
        def indexFromSheet(self, sheet, iLigne):
            #
            # Recherche les lettres de colonnes correspondante
            #
            self.list = []
            for iCol in range(1, sheet.max_column):
                if sheet.cell(iLigne, iCol).value != None:
                    self.append(sheet.cell(iLigne, iCol).value, iCol)
            
            if verbose > 5:    
                for i, val in enumerate(self.list):
                    print val.refCol, val.col
                    
        def getCol(self, refCol):
            for i, val in enumerate(self.list):
                if refCol == val.refCol:
                    return val.col
            print "RefCol non trouve"
            return ""
        
    def __init__(self, sheet, iLigneListCol):
        self.sheet = sheet
        self.firstLine = 1
        self.nbLigneVideMax = 100
        self.listCol = self.CLASS_indexCol()
        self.listCol.indexFromSheet(sheet, iLigneListCol)
        
    def searchLastLigneInColStr(self, refCol):
        return self.searchLastLigneInColNum(self, self.listCol.getCol(refCol))
        
    def searchLastLigneInColNum(self, iCol):
        continuer1 = True
        iLigne = self.firstLine
        cptLigneVide = 0
        ligneVide = 0
        while continuer1:
            if self.sheet.cell(iLigne, iCol).value == None:
                ligneVide = iLigne
                continuer2 = True
                while continuer2:
                    if self.sheet.cell(iLigne, iCol).value == None:
                        cptLigneVide += 1
                    else:
                        continuer2 = False
                    if cptLigneVide > self.nbLigneVideMax:
                        return ligneVide
                    iLigne += 1
            else:
                cptLigneVide = 0
                iLigne += 1

    def searchMaxInColStr(self, refCol):
        return self.searchMaxInColNum(self, self.listCol.getCol(refCol))
        
    def searchMaxInColNum(self, iCol):
        valMax = -1
        for iLigne in range(1, self.sheet.max_row):
            if self.sheet.cell(iLigne, iCol).value != None:
                if isinstance(self.sheet.cell(iLigne, iCol).value, basestring) == False:
                    if self.sheet.cell(iLigne, iCol).value > valMax:
                        valMax = self.sheet.cell(iLigne, iCol).value
                
        return valMax

    def searchValInColStr(self, refCol, val):
        return self.searchValInColNum(self, self.listCol.getCol(refCol), val)

    def searchValInColNum(self, iCol, val):
        listResult = []
        for iLigne in range(1, self.sheet.max_row):
            if self.sheet.cell(iLigne, iCol).value != None:
                if self.sheet.cell(iLigne, iCol).value > val:
                    listResult.append(iLigne)
        return listResult
    
def creatXlsx ():
    wb = Workbook()
    
    # grab the active worksheet
    ws = wb.active
    
    # Data can be assigned directly to cells
    ws['A1'] = 42
    
    # Rows can also be appended
    ws.append([1, 2, 3])
    
    # Python types will automatically be converted
    import datetime
    ws['A2'] = datetime.datetime.now()
    
    # Save the file
    wb.save("sample.xlsx")
    
if __name__ == "__main__":
    
    wb2 = load_workbook('sample.xlsx')
    out =  wb2.get_sheet_names()
    
    resultSheet = CLASS_resultSheet(wb2[out[1]], 1)

    print resultSheet.searchLastLigneInColStr("I_Date") 
    print resultSheet.searchLastLigneInColStr("I_SN_APP_AAAA") 
    
    print resultSheet.searchMaxInColStr("I_Date") 

    quit()
    
